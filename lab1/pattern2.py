import math
import openpyxl as xl
training_data_list = []
workbook = xl.load_workbook('input.xlsx')
sheet = workbook['Sheet1']
for row in range(1, sheet.max_row+1):

    height = sheet.cell(row, 1).value
    weight = sheet.cell(row, 2).value
    class_name = sheet.cell(row, 3).value
    data_dictionary = {'height' : height,
                       'weight' : weight,
                       'class' : class_name}
    training_data_list.append(data_dictionary)

for i in range(3):

    print("Enter the query")
    print("Height(inc) - Weight(Kg)")
    query_data = input()

    query_data2 = query_data.split(" ")
    test_height = int(query_data2[0])
    test_weight = int(query_data2[1])

    training_item = training_data_list[0]
    training_item_height = int(training_item['height'])
    training_item_weight = int(training_item['weight'])
    training_item_class = training_item['class']

    e_distance = math.sqrt((test_height-training_item_height)**2+
                           (test_weight-training_item_weight)**2)
    output_class_name = training_item_class

    for item in training_data_list:
        training_item_height = int(item['height'])
        training_item_weight = int(item['weight'])
        training_item_class = item['class']

        temp_e_distance = math.sqrt((test_height-training_item_height)**2+
                           (test_weight-training_item_weight)**2)
        if temp_e_distance < e_distance:
            e_distance = temp_e_distance
            output_class_name = training_item_class


    print("test data class is : ",output_class_name)